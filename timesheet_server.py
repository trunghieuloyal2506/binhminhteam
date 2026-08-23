"""
Cập nhật timesheet_server.py với:
1. Đơn vị NGÀY thay vì giờ (÷ 8.5)
2. Thêm header CORS đầy đủ hơn
"""
import json, calendar, io
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

CN='1F3864'; CL='D9E1F2'; CS='BDD7EE'; CU='FFDDC1'
CWE='D6DCE4'; CY='FFF2CC'; CO='F4B942'; CG='E2EFDA'; CW='FFFFFF'

def F(c): return PatternFill('solid', fgColor=c)
def Fn(bold=False, sz=9, col='000000', italic=False):
    return Font(name='Arial', bold=bold, size=sz, color=col, italic=italic)
def Al(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def Bd():
    t=Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def S(ws, r, ci, val=None, bold=False, sz=9, fg=None, fc='000000',
      h='center', v='center', wrap=False, bd=True, italic=False, nf=None):
    c=ws.cell(r,ci)
    if val is not None: c.value=val
    c.font=Fn(bold,sz,fc,italic); c.alignment=Al(h,v,wrap)
    if fg: c.fill=F(fg)
    if bd: c.border=Bd()
    if nf: c.number_format=nf
    return c

def M(ws,r1,c1,r2,c2):
    ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

def hrs_to_days(h, day_hrs=8.5):
    """Convert hours to days, rounded to 2 decimal places"""
    if not h: return 0
    return round(float(h)/day_hrs, 2)

def build_sheet(wb, member, month, year, tasks, leave_map, sheet_name=None):
    days=calendar.monthrange(year,month)[1]
    DAY=['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
    MN=['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

    ws=wb.create_sheet(sheet_name or member['name'][:28])
    TC=5+days

    ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=10
    ws.column_dimensions['C'].width=16; ws.column_dimensions['D'].width=7
    for i in range(1,days+1): ws.column_dimensions[get_column_letter(4+i)].width=4.2
    ws.column_dimensions[get_column_letter(TC)].width=7

    ws.row_dimensions[1].height=24
    ws.row_dimensions[8].height=20
    ws.row_dimensions[9].height=16

    # Title
    M(ws,1,1,1,TC); S(ws,1,1,'MONTHLY TIME SHEET',bold=True,sz=14,fg=CN,fc=CW,h='center')
    ws.row_dimensions[2].height=5

    # Info
    wd=sum(1 for dd in range(1,days+1) if date(year,month,dd).weekday()<5)
    est=round(wd*1.0,1)  # days (not hours)
    for r,(lbl,val) in [(3,('Full Name:',member['name'])),
                        (4,('Position:',member.get('role','') or 'QS Engineer')),
                        (5,('Month / Year:',f"{MN[month]} {year}"))]:
        ws.row_dimensions[r].height=17
        S(ws,r,1,lbl,bold=True,sz=10,fg=CL,h='left')
        M(ws,r,2,r,TC); c=ws.cell(r,2); c.value=val
        c.font=Fn(sz=10); c.alignment=Al(h='left'); c.fill=F(CW); c.border=Bd()

    ws.row_dimensions[6].height=17
    S(ws,6,1,'Working days:',bold=True,sz=10,fg=CL,h='left')
    S(ws,6,2,wd,bold=True,sz=10,fg=CW)
    S(ws,6,3,'Estimated Days:',bold=True,sz=10,fg=CL,h='left')
    S(ws,6,4,est,bold=True,sz=10,fg=CW)
    M(ws,6,5,6,TC); ws.cell(6,5).fill=F(CW); ws.cell(6,5).border=Bd()
    ws.row_dimensions[7].height=5

    # Day headers
    M(ws,8,1,9,1); S(ws,8,1,'Brief Description',bold=True,sz=9,fg=CN,fc=CW,h='center',wrap=True)
    M(ws,8,2,9,2); S(ws,8,2,'Code',bold=True,sz=9,fg=CN,fc=CW)
    M(ws,8,3,9,3); S(ws,8,3,'Package',bold=True,sz=9,fg=CN,fc=CW)
    M(ws,8,4,9,4); S(ws,8,4,'Task',bold=True,sz=9,fg=CN,fc=CW)

    for dd in range(1,days+1):
        ci=4+dd; wi=date(year,month,dd).weekday()
        bg=CS if wi==5 else (CU if wi==6 else CN)
        S(ws,8,ci,DAY[wi],bold=True,sz=8,fg=bg,fc=CW)
        S(ws,9,ci,dd,bold=True,sz=8,fg=bg,fc=CW)

    M(ws,8,TC,9,TC); S(ws,8,TC,'Total',bold=True,sz=9,fg=CO,fc=CW)

    # Section grouping
    pre  =[t for t in tasks if t.get('section')=='pre' or t.get('phase') in('cost','pre')]
    post =[t for t in tasks if t.get('section')=='post' or (t.get('phase') in('post','qs') and t not in pre)]
    other=[t for t in tasks if t not in pre and t not in post]
    cur=10

    def sec(row,lbl):
        ws.row_dimensions[row].height=16
        M(ws,row,1,row,TC); c=ws.cell(row,1)
        c.value=lbl; c.font=Fn(True,9,'1F3864',italic=True)
        c.fill=F(CL); c.alignment=Al(h='left'); c.border=Bd()
        return row+1

    def trow(row,t):
        ws.row_dimensions[row].height=15
        hrs=t.get('hours',{})
        S(ws,row,1,t.get('projName',''),sz=9,fg=CG,h='left')
        S(ws,row,2,t.get('projCode',''),sz=9,fg=CW,h='center')
        S(ws,row,3,t.get('package',''),sz=8,fg=CW,h='center')
        S(ws,row,4,t.get('taskCode',''),sz=9,fg=CW,h='center')
        tot=0
        for dd in range(1,days+1):
            ci=4+dd; wi=date(year,month,dd).weekday(); is_we=wi>=5
            hv=float(hrs.get(str(dd),0) or hrs.get(dd,0) or 0)
            dv=hrs_to_days(hv)  # convert to days
            bg=CWE if is_we else CW
            if dv:
                S(ws,row,ci,dv,sz=9,fg=bg,h='center',nf='0.00'); tot+=dv
            else:
                c=ws.cell(row,ci); c.fill=F(bg); c.border=Bd(); c.alignment=Al()
        S(ws,row,TC,round(tot,2) if tot else 0,bold=True,sz=9,fg=CY,h='center',nf='0.00')
        return row+1

    def erow(row):
        ws.row_dimensions[row].height=8
        for ci in range(1,TC+1):
            dd=ci-4; is_we=(ci>4)and(1<=dd<=days)and date(year,month,dd).weekday()>=5
            c=ws.cell(row,ci); c.fill=F(CWE if is_we and ci>4 else CW); c.border=Bd()
        return row+1

    if pre:
        cur=sec(cur,'Pre-Contract works')
        for t in pre: cur=trow(cur,t)
        cur=erow(cur)
    if post:
        cur=sec(cur,'Post-Contract works')
        for t in post: cur=trow(cur,t)
        cur=erow(cur)
    if other:
        cur=sec(cur,'Other works')
        for t in other: cur=trow(cur,t)
        cur=erow(cur)

    # Leave rows (in days)
    for lbl,code in [('Travelling - outside office','TT'),('Sick Leave','SL'),
                     ('Special Leave','SPL'),('Annual Leave','AL'),('Public Holidays','PL')]:
        lhrs=leave_map.get(code,{})
        t={'projName':lbl,'projCode':code,'package':'','taskCode':'',
           'section':'','phase':'',
           'hours':{str(k):v for k,v in lhrs.items()}}
        cur=trow(cur,t)

    # Total per day
    tr=cur; ws.row_dimensions[tr].height=18; ds=10; de=tr-1
    M(ws,tr,1,tr,4); S(ws,tr,1,'Total per day',bold=True,sz=9,fg=CN,fc=CW,h='left')
    for dd in range(1,days+1):
        ci=4+dd; cl=get_column_letter(ci); wi=date(year,month,dd).weekday()
        bg=CWE if wi>=5 else CN
        c=ws.cell(tr,ci); c.value=f'=SUM({cl}{ds}:{cl}{de})'
        c.font=Fn(True,9,CW); c.fill=F(bg); c.alignment=Al(); c.border=Bd(); c.number_format='0.00'
    c=ws.cell(tr,TC); c.value=f'=SUM({get_column_letter(TC)}{ds}:{get_column_letter(TC)}{de})'
    c.font=Fn(True,9,CW); c.fill=F(CO); c.alignment=Al(); c.border=Bd(); c.number_format='0.00'

    # Notes
    nr=tr+2; ws.row_dimensions[nr].height=16; M(ws,nr,1,nr,TC)
    ws.cell(nr,1).value='Notes (Unit: Days)'; ws.cell(nr,1).font=Fn(True,10)
    ws.cell(nr,1).alignment=Al(h='left')
    for i,n in enumerate([
        '1. Unit: Days (1 day = 8.5 hours). E.g. 0.5 = half day, 1.0 = full day.',
        '2. The time recorded against each day (except Saturdays and Sundays) shall equal 1.0 day.',
        '3. Please identify clearly the Pre-Contract or Post Contract Works'],1):
        r=nr+i; ws.row_dimensions[r].height=14; M(ws,r,1,r,TC)
        ws.cell(r,1).value=n; ws.cell(r,1).font=Fn(sz=9); ws.cell(r,1).alignment=Al(h='left')

    # Signatures
    sr=nr+6; ws.row_dimensions[sr].height=18; m1=4; m2=TC//2
    M(ws,sr,1,sr,m1);    S(ws,sr,1,'PREPARED BY',bold=True,sz=9,fg=CL)
    M(ws,sr,m1+1,sr,m2); S(ws,sr,m1+1,'REVIEWED BY',bold=True,sz=9,fg=CL)
    M(ws,sr,m2+1,sr,TC); S(ws,sr,m2+1,'CHECKED BY',bold=True,sz=9,fg=CL)
    name_r=sr+5; ws.row_dimensions[name_r].height=16
    M(ws,name_r,1,name_r,m1);    S(ws,name_r,1,member['name'],bold=True,sz=10,h='center',fg=CW)
    M(ws,name_r,m1+1,name_r,m2); S(ws,name_r,m1+1,'(Line Manager)',sz=9,h='center',fg=CW)
    M(ws,name_r,m2+1,name_r,TC); S(ws,name_r,m2+1,'Võ Thị Duyên - HR Dept',sz=9,h='center',fg=CW)

    ws.freeze_panes='E10'

def build_summary(wb, all_md, month, year):
    MN=['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    ws=wb.create_sheet('SUMMARY')

    # Collect all unique projects/leave rows
    proj_rows = {}  # key: (label, code_or_proj) -> {member_name: days}
    member_names = [md['member']['name'] for md in all_md]

    for md in all_md:
        name = md['member']['name']
        days_in_month = calendar.monthrange(year, month)[1]

        # Work tasks grouped by project
        for t in md['tasks']:
            proj_name = t.get('projName','?')
            phase = t.get('phase','')
            phase_label = {'cost':'Cost Plan','pre':'Pre-Tender','post':'Post-Tender','qs':'QS Service'}.get(phase,'')
            proj_code = t.get('projCode','')
            key = proj_name
            code_prefix = f"{proj_code}. " if proj_code and proj_code != proj_name else ""
            label = f"{code_prefix}{proj_name}" + (f" ({phase_label})" if phase_label else "")
            total_h = sum(float(v) for v in t.get('hours',{}).values() if v)
            total_d = hrs_to_days(total_h)
            if total_d > 0:
                if key not in proj_rows: proj_rows[key] = {'label': label, 'data': {}}
                proj_rows[key]['data'][name] = proj_rows[key]['data'].get(name, 0) + total_d

        # Leave rows
        LEAVE_LABELS = {'TT':'Travelling','SL':'Sick Leave','SPL':'Special Leave','AL':'Annual Leave','PL':'Public Holidays'}
        for code, lbl in LEAVE_LABELS.items():
            lhrs = md.get('leave_map', {}).get(code, {})
            total_h = sum(float(v) for v in lhrs.values() if v)
            total_d = hrs_to_days(total_h)
            if total_d > 0:
                if code not in proj_rows: proj_rows[code] = {'label': lbl, 'data': {}}
                proj_rows[code]['data'][name] = proj_rows[code]['data'].get(name, 0) + total_d

    # Column widths
    n_members = len(member_names)
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 26
    for i in range(n_members):
        ws.column_dimensions[get_column_letter(3+i)].width = 13
    ws.column_dimensions[get_column_letter(3+n_members)].width = 10

    # Title
    title_end = 2 + n_members + 1  # B + members + Total
    M(ws,1,1,1,title_end)
    c=ws.cell(1,1); c.value=f'TEAM TIMESHEET SUMMARY — {MN[month]} {year} (Unit: Days)'
    c.font=Fn(True,13,CW); c.fill=F(CN); c.alignment=Al(); ws.row_dimensions[1].height=24

    # Header row: STT | Dự án / Nghỉ | Name1 | Name2 | ... | Tổng
    ws.row_dimensions[3].height=20
    S(ws,3,1,'STT',bold=True,sz=10,fg=CN,fc=CW,h='center')
    S(ws,3,2,'Dự án / Loại nghỉ',bold=True,sz=10,fg=CN,fc=CW,h='left')
    for i, name in enumerate(member_names):
        short = name.split()[-1] if ' ' in name else name  # last name only
        S(ws,3,3+i,short,bold=True,sz=9,fg=CN,fc=CW,h='center',wrap=True)
    S(ws,3,3+n_members,'Tổng',bold=True,sz=10,fg=CO,fc=CW,h='center')

    # Data rows
    row = 4
    # Pre-Contract section header
    pre_projs = {k:v for k,v in proj_rows.items() if k not in ['TT','SL','SPL','AL','PL']}
    leave_projs = {k:v for k,v in proj_rows.items() if k in ['TT','SL','SPL','AL','PL']}

    def write_section_header(ws, row, label):
        M(ws,row,1,row,3+n_members)
        c=ws.cell(row,1); c.value=label
        c.font=Fn(True,9,'1F3864',italic=True)
        c.fill=F(CL); c.alignment=Al(h='left'); c.border=Bd()
        ws.row_dimensions[row].height=16
        return row+1

    if pre_projs:
        row = write_section_header(ws, row, 'Work')
        for idx,(key,info) in enumerate(pre_projs.items(),1):
            ws.row_dimensions[row].height=15
            bg = 'F5F5F5' if idx%2==0 else CW
            S(ws,row,1,idx,sz=9,fg=bg,h='center')
            S(ws,row,2,info['label'],sz=9,fg=bg,h='left')
            row_total = 0
            for i,name in enumerate(member_names):
                val = round(info['data'].get(name,0), 2)
                c=ws.cell(row,3+i); c.value=val if val else ''
                c.font=Fn(sz=9); c.alignment=Al(h='center'); c.border=Bd()
                c.fill=F(bg); c.number_format='0.00'
                row_total += val
            S(ws,row,3+n_members,round(row_total,2) if row_total else '',bold=True,sz=9,fg=CY,h='center',nf='0.00')
            row+=1

    if leave_projs:
        row = write_section_header(ws, row, 'Leave / Nghỉ')
        for idx,(key,info) in enumerate(leave_projs.items(),1):
            ws.row_dimensions[row].height=15
            bg = 'FFF9C4' if idx%2==0 else 'FFFDE7'
            S(ws,row,1,idx,sz=9,fg=bg,h='center')
            S(ws,row,2,info['label'],sz=9,fg=bg,h='left')
            row_total=0
            for i,name in enumerate(member_names):
                val=round(info['data'].get(name,0),2)
                c=ws.cell(row,3+i); c.value=val if val else ''
                c.font=Fn(sz=9); c.alignment=Al(h='center'); c.border=Bd()
                c.fill=F(bg); c.number_format='0.00'
                row_total+=val
            S(ws,row,3+n_members,round(row_total,2) if row_total else '',bold=True,sz=9,fg=CY,h='center',nf='0.00')
            row+=1

    # Total row
    ws.row_dimensions[row].height=20
    M(ws,row,1,row,2); S(ws,row,1,'TỔNG CỘNG',bold=True,sz=11,fg=CY,h='center')
    grand=0
    for i,name in enumerate(member_names):
        total_d=round(sum(
            info['data'].get(name,0) for info in proj_rows.values()
        ),2)
        grand+=total_d
        c=ws.cell(row,3+i); c.value=total_d if total_d else 0
        c.font=Fn(True,11); c.fill=F(CY); c.alignment=Al(h='center'); c.border=Bd()
        c.number_format='0.00'
        # Color by workload %
        wd=all_md[i]['working_days']
        pct=round(total_d/wd*100,1) if wd>0 else 0
        c.font=Fn(True,11,'EF4444' if pct>120 else ('F97316' if pct>110 else ('F59E0B' if pct>100 else '1F3864')))
    S(ws,row,3+n_members,round(grand,2),bold=True,sz=11,fg=CY,h='center',nf='0.00')

    # % row
    row+=1; ws.row_dimensions[row].height=16
    M(ws,row,1,row,2); S(ws,row,1,'% Công suất',bold=True,sz=10,fg=CL,h='center')
    for i,md in enumerate(all_md):
        total_d=round(sum(info['data'].get(md['member']['name'],0) for info in proj_rows.values()),2)
        wd=md['working_days']
        pct=round(total_d/wd*100,1) if wd>0 else 0
        col='EF4444' if pct>120 else ('F97316' if pct>110 else ('F59E0B' if pct>100 else '15803D'))
        S(ws,row,3+i,f'{pct}%',bold=True,sz=10,fg=CL,fc=col,h='center')
    S(ws,row,3+n_members,'',fg=CL)

    ws.freeze_panes='C4'


class Handler(BaseHTTPRequestHandler):
    def log_message(self,format,*args):
        print(f"[Server] {args[0]} {args[1]} {args[2]}")

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors(); self.end_headers()

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin','*')
        self.send_header('Access-Control-Allow-Methods','POST,OPTIONS,GET')
        self.send_header('Access-Control-Allow-Headers','Content-Type,Accept')

    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type','text/plain')
        self._cors(); self.end_headers()
        self.wfile.write(b'VinaQS Timesheet Server OK')

    def do_POST(self):
        try:
            length=int(self.headers.get('Content-Length',0))
            body=self.rfile.read(length)
            data=json.loads(body)
            month=int(data['month']); year=int(data['year'])
            members_data=data['members']
            days_in_month=calendar.monthrange(year,month)[1]
            wd=sum(1 for dd in range(1,days_in_month+1) if date(year,month,dd).weekday()<5)

            wb=Workbook(); wb.remove(wb.active)
            all_md=[]
            for md in members_data:
                md['working_days']=wd; all_md.append(md)
                name=md['member']['name'][:28].replace('/','_')
                build_sheet(wb,md['member'],month,year,md['tasks'],md.get('leave_map',{}),name)
            build_summary(wb,all_md,month,year)

            buf=io.BytesIO(); wb.save(buf); buf.seek(0); xlsx=buf.read()
            self.send_response(200)
            self.send_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition',f'attachment; filename="TimeSheet_{month:02d}_{year}_Team.xlsx"')
            self.send_header('Content-Length',str(len(xlsx)))
            self._cors(); self.end_headers()
            self.wfile.write(xlsx)
            print(f"[Server] ✅ Exported {len(members_data)} members")
        except Exception as e:
            import traceback; traceback.print_exc()
            self.send_response(500); self.send_header('Content-Type','application/json')
            self._cors(); self.end_headers()
            self.wfile.write(json.dumps({'error':str(e)}).encode())

if __name__=='__main__':
    print("="*50)
    print("  VinaQS Timesheet Server (Unit: Days)")
    print("  http://localhost:5000")
    print("  Để dừng: Ctrl+C")
    print("="*50)
    HTTPServer(('localhost',5000),Handler).serve_forever()
